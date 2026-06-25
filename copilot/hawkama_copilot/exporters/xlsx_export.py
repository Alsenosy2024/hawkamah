"""XLSX export (openpyxl).

Every Markdown table becomes its own RTL worksheet (named after the nearest
preceding heading), plus an "المحتوى" outline sheet listing headings & paragraphs.
Useful for the matrix-heavy deliverables (RACI, DoA, KPIs, risk register, gap
matrix) where the operator wants the tables as real spreadsheets.
"""

from __future__ import annotations

import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .markdown_ast import Block, parse_markdown

NAVY_FILL = PatternFill("solid", fgColor="1A3557")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Almarai")
BODY_FONT = Font(name="Almarai")
RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True, readingOrder=2)


def export_xlsx(markdown: str, title: str) -> bytes:
    wb = openpyxl.Workbook()
    blocks = parse_markdown(markdown)

    outline = wb.active
    outline.title = _sheet_name("المحتوى")
    outline.sheet_view.rightToLeft = True
    outline.append(["العنوان/الفقرة", "المستوى"])
    _style_header(outline, 2)

    current_heading = title
    used_names: set[str] = {outline.title}
    table_idx = 0

    for b in blocks:
        if b.type == "heading":
            current_heading = b.text
            outline.append([b.text, f"H{b.level}"])
            _style_body_row(outline, outline.max_row, 2)
        elif b.type in ("paragraph", "bullet", "ordered", "quote"):
            outline.append([b.text, b.type])
            _style_body_row(outline, outline.max_row, 2)
        elif b.type == "table":
            table_idx += 1
            name = _unique(_sheet_name(current_heading or f"جدول {table_idx}"), used_names)
            used_names.add(name)
            ws = wb.create_sheet(name)
            ws.sheet_view.rightToLeft = True
            _write_table(ws, b)

    _autosize(outline)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_table(ws, b: Block) -> None:
    cols = max(len(b.headers), max((len(r) for r in b.rows), default=0)) or 1
    ws.append(b.headers + [""] * (cols - len(b.headers)))
    _style_header(ws, cols)
    for r in b.rows:
        ws.append(r + [""] * (cols - len(r)))
        _style_body_row(ws, ws.max_row, cols)
    _autosize(ws)


def _style_header(ws, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = RIGHT
    ws.freeze_panes = "A2"


def _style_body_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.alignment = RIGHT


def _autosize(ws) -> None:
    for col in ws.columns:
        width = 12
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                width = max(width, min(60, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _sheet_name(name: str) -> str:
    # Excel sheet names: <=31 chars, no []:*?/\
    clean = re.sub(r"[\[\]:*?/\\]", " ", name).strip() or "ورقة"
    return clean[:31]


def _unique(name: str, used: set[str]) -> str:
    if name not in used:
        return name
    for i in range(2, 100):
        cand = f"{name[:28]}_{i}"
        if cand not in used:
            return cand
    return name[:28] + "_x"
